from __future__ import annotations
import os
from typing import List, Optional, TYPE_CHECKING
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from uuid import uuid4
from pandas import DataFrame, concat
if TYPE_CHECKING:
    from geoprob_pipe.utils.df_validation.column import ColumnValidation


# TODO
#  - Nice to have   Export to Excel: message if validation is stopped early;


class DataFrameValidation:

    def __init__(
            self, df: DataFrame, label: str,
            columns_validations: List[ColumnValidation], required_columns: List[str] = None,
    ):
        self.label: str = label
        self.df: DataFrame = df
        self.required_columns: Optional[List[str]] = required_columns
        self.columns_validations: List[ColumnValidation] = columns_validations

        # Placeholders
        self.df_failures: Optional[DataFrame] = None

    def _run_required_columns(self) -> DataFrame:
        failure_rows = []
        for required_column in self.required_columns:
            if required_column not in self.df.columns:
                failure_rows.append({
                    "failure_msg": f"Column '{required_column}' is missing in the provided DataFrame.",
                    "dataframe": self.label,
                })
        return DataFrame(failure_rows)

    def _run_column_validation(self) -> DataFrame:
        failures = []

        for col_val in self.columns_validations:
            for req in col_val.requirements:

                # Perform validation
                df_invalid = req.validate(self.df, col_val.column_name)
                if not df_invalid.empty:
                    failures.append(df_invalid)

                # Stop on failure?
                if req.stop_validation_on_failure and df_invalid.__len__() > 0:
                    return concat(failures)

        if failures:
            return concat(failures)
        return DataFrame()

    def run(self):

        df = self._run_required_columns()
        if df.__len__() > 0:
            df = df.reset_index(drop=True)
            self.df_failures = df
            return

        df = self._run_column_validation()
        if df.__len__() > 0:
            df = df.reset_index(drop=True)
            self.df_failures = df
            return

    def to_excel(self, directory: str):
        df: DataFrame = self.df_failures.copy(deep=True)

        # Determine export path
        filename = "failed_validation_requirements"
        os.makedirs(directory, exist_ok=True)
        export_path = os.path.join(directory, f"{filename}.xlsx")
        if os.path.exists(export_path):
            export_path = os.path.join(directory, f"{filename}_{uuid4().__str__()}.xlsx")

        # Export
        df.to_excel(export_path)

        # Color cell
        if "column" not in df.columns:
            return
        wb = load_workbook(export_path)
        ws = wb.active
        red_fill = PatternFill(start_color="FFDA9694", end_color="FFDA9694", fill_type="solid")
        for index, (_, row) in enumerate(df.iterrows(), start=2):
            scope_col_idx = df.columns.get_loc(row['column']) + 2
            ws.cell(row=index, column=scope_col_idx).fill = red_fill
        wb.save(export_path)
