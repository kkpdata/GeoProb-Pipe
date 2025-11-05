from __future__ import annotations
import os
import sqlite3
import shutil
from openpyxl import load_workbook
from pandas import ExcelWriter, DataFrame
from geopandas import read_file, GeoDataFrame
from datetime import datetime
import importlib.resources
from typing import TYPE_CHECKING
from geoprob_pipe.calculations.system_calculations import SYSTEM_CALCULATION_MAPPER
from geoprob_pipe.utils.validation_messages import BColors
if TYPE_CHECKING:
    from geoprob_pipe.pre_processing.cmd import ApplicationSettings
    from geoprob_pipe.pre_processing.parameter_input.input_parameter_tables import InputParameterTables


def join_vak_naam(df: DataFrame, app_settings: ApplicationSettings, ) -> DataFrame:
    gdf_vakindeling: GeoDataFrame = read_file(app_settings.geopackage_filepath, layer="vakindeling")
    gdf_vakindeling = gdf_vakindeling.rename(columns={"naam": "vak_naam", "id": "vak_id"})
    df = df.merge(gdf_vakindeling[["vak_naam", "vak_id"]], left_on="vak_id", right_on="vak_id")
    return df[["vak_id", "vak_naam", "naam", "kans"]]


def export_input_parameter_tables(app_settings: ApplicationSettings, tables: InputParameterTables):

    # Copy template to workspace
    dst_dir = os.path.join(app_settings.workspace_dir, "parameter_input_process")
    os.makedirs(dst_dir, exist_ok=True)
    dst_path = os.path.join(dst_dir, "input_parameters_template.xlsx")
    dst_path = dst_path.replace("_template.xlsx", f".xlsx")
    if os.path.exists(dst_path):
        datetime_stamp = datetime.now().strftime("%Y-%m-%d_%H%M%S")
        dst_path = dst_path.replace(".xlsx", f"_{datetime_stamp}.xlsx")
    with importlib.resources.path(
            'geoprob_pipe.pre_processing.parameter_input', 'parameter_input_template.xlsx'
    ) as src_path:
        shutil.copy2(src=src_path, dst=dst_path)

    # Fill 'Model parameters'
    conn = sqlite3.connect(app_settings.geopackage_filepath)
    cursor = conn.cursor()
    cursor.execute("""
        SELECT geoprob_pipe_metadata."values" 
        FROM geoprob_pipe_metadata 
        WHERE metadata_type='geohydrologisch_model';
    """)
    result = cursor.fetchone()
    if not result:
        raise ValueError
    model_string = result[0]
    conn.close()
    df_dummy_data = DataFrame(SYSTEM_CALCULATION_MAPPER[model_string]['dummy_invoer'])
    df_dummy_data = df_dummy_data.sort_values(by=["name"])
    df_model_parameters = df_dummy_data[["name", "description", "remark", "unit"]].copy()
    with ExcelWriter(dst_path, engine="openpyxl", mode="a", if_sheet_exists="overlay") as writer:
        df_model_parameters.to_excel(
            writer, sheet_name="Geohydrologisch model", index=False, header=False, startrow=6, startcol=0)
    # TODO: Kolom 'invullen' is nog onduidelijk en niet gevuld.
    # Geohydrologisch model in titel
    wb = load_workbook(dst_path)
    ws = wb["Geohydrologisch model"]
    ws["A1"] = f"Geohydrologisch model '{SYSTEM_CALCULATION_MAPPER[model_string]['label']}'"
    wb.save(dst_path)

    # Fill 'Scenario invoer'
    df_scenario_invoer = tables.df_scenario_invoer
    df_scenario_invoer = join_vak_naam(df=df_scenario_invoer, app_settings=app_settings)
    with ExcelWriter(dst_path, engine="openpyxl", mode="a", if_sheet_exists="overlay") as writer:
        df_scenario_invoer.to_excel(
            writer, sheet_name="Scenario invoer", index=False, header=False, startrow=3, startcol=0)

    # Fill 'Parameter invoer'
    df_parameter_invoer = tables.df_parameter_invoer
    df_parameter_invoer = df_parameter_invoer[[
        "parameter", "scope", "scope_referentie", "ondergrondscenario_naam", "distribution_type", "mean", "variation",
        "deviation", "minimum", "maximum", "fragility_values_ref", "bronnen", "opmerking"]].copy()
    with ExcelWriter(dst_path, engine="openpyxl", mode="a", if_sheet_exists="overlay") as writer:
        df_parameter_invoer.to_excel(
            writer, sheet_name="Parameter invoer", index=False, header=False, startrow=4, startcol=0)

    # Fill 'Fragility values'
    df_fragility_values_invoer = tables.df_fragility_values_invoer
    df_fragility_values_invoer = df_fragility_values_invoer[["fragility_values_ref", "waarde", "kans"]].copy()
    with ExcelWriter(dst_path, engine="openpyxl", mode="a", if_sheet_exists="overlay") as writer:
        df_fragility_values_invoer.to_excel(
            writer, sheet_name="Fragility values", index=False, header=False, startrow=4, startcol=0)

    print(f"{BColors.UNDERLINE}Exporteren van invoer tabellen compleet:\n"
          f"{dst_path}{BColors.ENDC}")
